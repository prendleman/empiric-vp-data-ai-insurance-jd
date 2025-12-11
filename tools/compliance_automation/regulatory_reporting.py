"""
Regulatory Reporting Automation

Automates generation of regulatory reports for insurance compliance.
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import json
import os

class RegulatoryReporter:
    """Generates regulatory reports for insurance compliance."""
    
    def __init__(self):
        """Initialize regulatory reporter."""
        self.report_templates = {
            'naic': self._naic_template,
            'state_annual': self._state_annual_template,
            'quarterly': self._quarterly_template
        }
    
    def generate_naic_report(self, data_path=None, policy_df=None, output_path='naic_report.xlsx'):
        """
        Generate NAIC (National Association of Insurance Commissioners) report.
        
        Args:
            data_path: Path to policy data CSV
            policy_df: DataFrame with policy data
            output_path: Path to save report
        """
        # Load data
        if data_path:
            df = pd.read_csv(data_path)
        elif policy_df is not None:
            df = policy_df.copy()
        else:
            raise ValueError("Must provide data_path or policy_df")
        
        # Generate report sections
        report_data = {
            'summary': self._generate_summary_section(df),
            'policy_counts': self._generate_policy_counts(df),
            'premium_analysis': self._generate_premium_analysis(df),
            'lapse_analysis': self._generate_lapse_analysis(df),
            'state_breakdown': self._generate_state_breakdown(df)
        }
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "NAIC Report"
        
        # Write header
        ws['A1'] = 'NAIC Regulatory Report'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f'Report Period: {datetime.now().strftime("%Y-%m-%d")}'
        
        # Write summary
        row = 4
        ws[f'A{row}'] = 'Summary Statistics'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for key, value in report_data['summary'].items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            row += 1
        
        # Write policy counts
        row += 2
        ws[f'A{row}'] = 'Policy Counts by Type'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Policy Type'
        ws[f'B{row}'] = 'Count'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        
        for policy_type, count in report_data['policy_counts'].items():
            ws[f'A{row}'] = policy_type
            ws[f'B{row}'] = count
            row += 1
        
        # Save workbook
        wb.save(output_path)
        print(f"NAIC report saved to {output_path}")
        
        return report_data
    
    def _generate_summary_section(self, df):
        """Generate summary statistics section."""
        return {
            'Total Policies': len(df),
            'Active Policies': len(df[df['status'] == 'Active']) if 'status' in df.columns else len(df),
            'Total Annual Premium': f"${df['annual_premium'].sum():,.2f}" if 'annual_premium' in df.columns else 'N/A',
            'Average Premium': f"${df['annual_premium'].mean():,.2f}" if 'annual_premium' in df.columns else 'N/A',
            'Report Date': datetime.now().strftime('%Y-%m-%d')
        }
    
    def _generate_policy_counts(self, df):
        """Generate policy counts by type."""
        if 'policy_type' in df.columns:
            return df['policy_type'].value_counts().to_dict()
        return {'Total': len(df)}
    
    def _generate_premium_analysis(self, df):
        """Generate premium analysis."""
        if 'annual_premium' not in df.columns:
            return {}
        
        return {
            'total_premium': df['annual_premium'].sum(),
            'average_premium': df['annual_premium'].mean(),
            'median_premium': df['annual_premium'].median(),
            'min_premium': df['annual_premium'].min(),
            'max_premium': df['annual_premium'].max()
        }
    
    def _generate_lapse_analysis(self, df):
        """Generate lapse analysis."""
        if 'status' not in df.columns:
            return {}
        
        total = len(df)
        lapsed = len(df[df['status'] == 'Lapsed'])
        
        return {
            'total_policies': total,
            'lapsed_policies': lapsed,
            'lapse_rate': lapsed / total if total > 0 else 0
        }
    
    def _generate_state_breakdown(self, df):
        """Generate state-by-state breakdown."""
        if 'state' not in df.columns:
            return {}
        
        state_summary = df.groupby('state').agg({
            'annual_premium': 'sum' if 'annual_premium' in df.columns else 'count',
            'policy_id': 'count' if 'policy_id' in df.columns else 'count'
        }).to_dict()
        
        return state_summary
    
    def _naic_template(self):
        """NAIC report template structure."""
        return {
            'sections': [
                'Summary Statistics',
                'Policy Counts by Type',
                'Premium Analysis',
                'Lapse Analysis',
                'State Breakdown'
            ]
        }
    
    def _state_annual_template(self):
        """State annual report template."""
        return {
            'sections': [
                'Executive Summary',
                'Policy Statistics',
                'Financial Summary',
                'Compliance Checklist'
            ]
        }
    
    def _quarterly_template(self):
        """Quarterly report template."""
        return {
            'sections': [
                'Quarter Summary',
                'Policy Activity',
                'Financial Metrics',
                'Compliance Status'
            ]
        }
    
    def generate_state_report(self, state, data_path=None, policy_df=None, output_path=None):
        """Generate state-specific regulatory report."""
        if output_path is None:
            output_path = f'{state}_annual_report.xlsx'
        
        # Similar structure to NAIC report but state-specific
        return self.generate_naic_report(data_path, policy_df, output_path)
    
    def get_compliance_deadlines(self, report_type='naic'):
        """Get compliance deadlines for regulatory reports."""
        current_year = datetime.now().year
        
        deadlines = {
            'naic': {
                'annual_report': datetime(current_year + 1, 3, 1),
                'quarterly_report_q1': datetime(current_year, 5, 15),
                'quarterly_report_q2': datetime(current_year, 8, 15),
                'quarterly_report_q3': datetime(current_year, 11, 15),
                'quarterly_report_q4': datetime(current_year + 1, 2, 15)
            },
            'state': {
                'annual_report': datetime(current_year + 1, 3, 31),
                'quarterly_report': datetime(current_year, 4, 30)
            }
        }
        
        return deadlines.get(report_type, {})

